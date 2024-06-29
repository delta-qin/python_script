
from bs4 import BeautifulSoup
from openpyxl import Workbook
import os

from openpyxl.reader.excel import load_workbook


def extract_content_from_html(html_file):
    with open(html_file, 'r', encoding='utf-8') as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, 'html.parser')
    entry_content = soup.find('div', class_='entry-content')

    return entry_content


def write_to_excel(content, excel_file):
    if not os.path.exists(excel_file):
        wb = Workbook()
    else:
        wb = load_workbook(excel_file)
    ws = wb.active
    ws.append([content])

    wb.save(excel_file)


def remove_something(content):
    index = content.find('热门推荐:')
    if index != -1:
        return content[:index]
    return content

# html_file = '/Users/deltaqin/Downloads/www.aihub.cn/tools/chatbot/agentgpt/index.html'
excel_file = 'aihub_output.xlsx'

start_dir = '/Users/deltaqin/Downloads/www.aihub.cn/tools'
for entry in os.scandir(start_dir):
    if entry.is_dir():
        for sub_entry in os.scandir(entry.path):
            if sub_entry.is_dir():
                for file in os.listdir(sub_entry.path):
                    if file.endswith('.html'):
                        html_file = os.path.join(sub_entry.path, file)
                        content = extract_content_from_html(html_file)
                        if content:
                            useful_content = content.get_text(separator='\n', strip=True)
                            write_to_excel(remove_something(useful_content), excel_file)
                        else:
                            print(f'No content found in {html_file}')


# content = extract_content_from_html(html_file)
# print(content)
# print(content.get_text(separator='\n', strip=True))
# print(content)
# write_to_excel(content, excel_file)
