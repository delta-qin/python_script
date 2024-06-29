"""
从 Excel 中读取数据，调用 OpenAI 生成JSON，调用gen_sql生成SQL，输出到文件，每个 sql 语句一行
"""
import os
import json
from OpenAI import simplify_text
from gen_sql import generate_sql_from_json
from openpyxl import load_workbook

from openpyxl import Workbook
import re
from ast import literal_eval
from openpyxl import Workbook
from openpyxl import load_workbook

def save_error_line_to_excel(data):
    try:
        # 尝试加载已存在的工作簿
        wb = load_workbook("logo_error.xlsx")
    except FileNotFoundError:
        # 如果文件不存在，则创建一个新的工作簿
        wb = Workbook()

    # 选择活动工作表
    ws = wb.active
    # 写入数据
    ws.append(data)
    # 保存工作簿
    wb.save("logo_error.xlsx")


def read_excel_rows(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    # 获取表格的行数和列数
    max_row = ws.max_row
    max_column = ws.max_column

    data = []
    # 读取每一行的数据
    for row in range(1, max_row + 1):
        row_data = []
        for col in range(1, max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            row_data.append(cell_value)
        data.append(row_data)
    return data

def content_to_sql(line):
    json_str = simplify_text(line)
    # 直接传递字符串就可以
    (sql, id_str, name, tags, category) = generate_sql_from_json(json_str)
    print(sql)
    with open('2024-04-08-logocom.sql', 'a') as file:
        file.write(sql + '\n')

    # 分类和标签写到Excel
    excel_file = "2024-04-08-logocom-app-tags.xlsx"
    if not os.path.exists(excel_file):
        wb = Workbook()
    else:
        wb = load_workbook(excel_file)
    ws = wb.active

    ws.append([id_str, name,  str(tags), category])
    wb.save(excel_file)

def inner_parse_sql_fields(sql):
    # 使用正则表达式匹配字段名和值
    fields_match = re.search(r"INSERT INTO `.*`\.`.*` \((.*)\) VALUES", sql)
    values_match = re.search(r"VALUES\s*\((.*)\);", sql)

    if fields_match and values_match:
        # 分割字段名
        fields = [field.strip("` ") for field in fields_match.group(1).split(',')]

        # 提取VALUES部分的值，考虑到值中可能有逗号，使用ast.literal_eval来解析
        values_str = f"[{values_match.group(1)}]"
        try:
            values = literal_eval(values_str)
        except Exception as e:
            print(f"Error parsing values: {e}")
            print(f"Values string: {values_str}")
            return None

        if len(fields) != len(values):
            print("The number of fields does not match the number of values.")
            return None

        # 返回字段和值的元组列表
        return list(zip(fields, values))
    else:
        print("Failed to match SQL pattern.")
        return None

def parse_sql_file(file_path):
    # 用于存储解析后的数据
    data = []

    # 打开SQL文件
    with open(file_path, 'r') as file:
        lines = file.readlines()
        # 按照每两行一组处理
        for i in range(0, len(lines), 2):
            sql_line = lines[i].strip() + lines[i + 1].strip()
            # 判断是否为SQL语句的行，以INSERT开头
            if sql_line.startswith("INSERT"):
                res = inner_parse_sql_fields(sql_line)
                if res:
                    name = res[0][1]
                    description = res[2][1]
                    application_id = res[5][1]
                    highlights = res[6][1]
                    faq = res[12][1]
                    # 删除字符串中的 \n
                    faq  = faq.replace('\n', '')
                    parsed_data = json.loads(faq)
                    new_string = ''
                    for item in parsed_data:
                        q = item.get('question', '')
                        a = item.get('answer', '')
                        new_string += f"{q} {a} "
                    # print(name, description, application_id, highlights,  new_string)
                    arr = [name, description, application_id, highlights, new_string]
                    data.append(arr)

    return data

def sql_to_tag_category():
    # 读取 SQL 文件
    parsed_data = parse_sql_file('2024-04-08-logocom.sql')
    # 拼接每一行内容。调用openai获取分类，写入Excel 200 798
    for line in parsed_data[697:798]:
        content = "名字：" + line[0] + "简介：" + line[1] + "亮点：" + line[3] + "FAQ：" + line[4]
        print(content)
        # 调用大模型转换为 sql
        json_str = simplify_text(content)
        # 直接传递字符串就可以
        res = json.loads(json_str)
        category = res.get("category", "")
        tags = res.get("tags", [])

        # 写到新的excel文件
        excel_file = "2024-04-08-logocom-app-tags.xlsx"
        if not os.path.exists(excel_file):
            wb = Workbook()
        else:
            wb = load_workbook(excel_file)
        ws = wb.active

        ws.append([line[2],  line[0] ,  str(tags), category])
        wb.save(excel_file)



def process():
    excel_file = 'output_logocom.xlsx'
    start = 0
    end = 963
    for line in read_excel_rows(excel_file)[start:end]:
        # 异常处理，保证剩余的可以继续执行，打印相关的错误信息
        try:
            content = "名字：" + line[0] + "官网：" + line[1] + "简介：" + line[2].replace('\n', '')
            print(content)
            # 调用大模型转换为sql
            content_to_sql(content)
        except Exception as e:
            print(e)
            print('ERROR 出错的行：', line)
            # 写到新的excel文件
            save_error_line_to_excel(line)


if __name__ == '__main__':
    # 解析 Excel 内容获取应用信息到sql
    # process()

    # 从sql文件中获取分类和标签
    # sql_to_tag_category()

    # 测试函数
    result = sql_to_tag_category()
