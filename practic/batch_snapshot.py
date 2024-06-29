from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from PIL import Image
from io import BytesIO
from openpyxl import Workbook
import os
from openpyxl.reader.excel import load_workbook

proxy_server = "127.0.0.1:7890"


def parse_sql_file(file_path):
    # 用于存储解析后的数据
    parsed_data = []

    # 打开SQL文件
    with open(file_path, 'r') as file:
        lines = file.readlines()
        # 按照每两行一组处理
        for i in range(0, len(lines), 2):
            sql_line = lines[i].strip() + lines[i + 1].strip()
            # 判断是否为SQL语句的行，以INSERT开头
            if sql_line.startswith("INSERT"):
                # 提取出每条SQL语句中的VALUES部分
                values_part = sql_line.split("VALUES")[1].strip("();")
                # 将VALUES部分按照逗号分割成列表
                values_list = values_part.split("', '")
                # 获取link和application_id的值，并添加到解析后的数据中
                link = values_list[3].strip("'")
                application_id = values_list[5].strip("'")
                # print(link, application_id)
                parsed_data.append([link, application_id])

    return parsed_data


def capture_screenshot_with_macbook(url, output_file):

    # Chromedriver的路径
    chromedriver_path = r"/Users/deltaqin/Downloads/chromedriver-mac-arm64/chromedriver"

    # 设置Chrome选项
    options = webdriver.ChromeOptions()
    options.add_experimental_option('detach', True)

    # 设置WebDriver服务
    service = Service(chromedriver_path)

    # 设置Chrome选项
    options.add_argument("--headless")  # 无头模式，不打开浏览器窗口
    options.add_argument('--proxy-server=%s' % proxy_server)

    # 启动Chrome浏览器
    driver = webdriver.Chrome(service=service, options=options)

    # 设置窗口大小为MacBook 16"分辨率
    driver.set_window_size(1440, 900)

    try:
        # 打开网页
        driver.get(url)

        # 截取网页截图
        screenshot = driver.get_screenshot_as_png()
        screenshot_image = Image.open(BytesIO(screenshot))

        # 加载Macbook设备外壳图片
        macbook_frame = Image.open("16.webp")  # 替换为你的Macbook外壳图片路径
        macbook_frame = macbook_frame.convert("RGBA")  # 确保图片有alpha通道

        # 调整网页截图大小以适应Macbook外壳指定的区域
        frame_width = 1293 - 144  # 框的宽度
        frame_height = 845 - 105  # 框的高度
        screenshot_image = screenshot_image.resize((frame_width, frame_height))

        # 创建新的图像对象，将网页截图和Macbook外壳叠加
        final_image = Image.new("RGBA", macbook_frame.size)
        final_image.paste(screenshot_image, (144, 105))
        final_image.alpha_composite(macbook_frame)

        # 保存最终的图像
        final_image.save(output_file)
    except Exception as e:
        print(e)
        print("Error:", link, application_id)
        # 如果图片文件存在，删除
        if os.path.exists(f"./img/{application_id}.png"):
            os.remove(f"./img/{application_id}.png")
        # 将异常的链接和ID写到Excel中
        save_error_link_to_excel(link, application_id)
    finally:
        # 关闭浏览器
        driver.quit()


def save_error_link_to_excel(link, application_id):
    excel_file = "png_error_links.xlsx"
    if not os.path.exists(excel_file):
        wb = Workbook()
    else:
        wb = load_workbook(excel_file)
    ws = wb.active
    ws.append([link, application_id])
    wb.save(excel_file)



if __name__ == '__main__':
    # 使用示例
    file_path = "2024-04-08-logocom.sql"
    parsed_data = parse_sql_file(file_path)
    # print(parsed_data[:500])
    for link, application_id in parsed_data[37:]:
        # 需要针对异常的链接进行处理，捕获之后将对应的ID和链接写到Excel中
        # 检测文件是否存在，如果存在则跳过
        if os.path.exists(f"./img/{application_id}.png"):
            continue
        capture_screenshot_with_macbook(link, f"./img/{application_id}.png")


        print(link, application_id)
        capture_screenshot_with_macbook(link, f"./img/{application_id}.png")

    # # 读取Excel
    # # 使用示例
    # url = "https://xbfox.com/"
    # output_file = "screenshot_with_macbook.png"
    # capture_screenshot_with_macbook(url, output_file)


