from bs4 import BeautifulSoup
from openpyxl import Workbook
import os
from urllib.parse import urlparse, parse_qs, urlunparse

from openpyxl.reader.excel import load_workbook


def extract_clean_url(soup):
    a_tags = soup.find_all('a', class_='group tw-s5xdrg')

    clean_urls = []
    for a_tag in a_tags:
        href = a_tag.get('href')
        if href:
            parsed_url = urlparse(href)
            # 直接截取 ? 后面的字符串删除
            clean_url = parsed_url._replace(query='').geturl()
            clean_urls.append(clean_url)

    return clean_urls


def extract_content_from_html(html_file):
    with open(html_file, 'r', encoding='utf-8') as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, 'html.parser')
    span_tag = soup.find('span', class_='css-643bw8 e1r4ce3211')
    for sup_tag in span_tag.find_all('sup'):
        sup_tag.decompose()
    name = span_tag.get_text(strip=True)

    res = extract_clean_url(soup)
    url = res[0]
    div_content = soup.find('div', class_='css-1b2vsf e1r4ce326').text.strip()

    # 删除 div_content 字符串开头的 简介：
    index = div_content.find('简介：')
    if index != -1:
        div_content = div_content[index + 3:]

    return (name, url , div_content)


def write_to_excel(name, url , div_content, excel_file):
    if not os.path.exists(excel_file):
        wb = Workbook()
    else:
        wb =load_workbook(excel_file)
    ws = wb.active
    ws.append([name, url, div_content])

    wb.save(excel_file)


# html_file = '/Users/deltaqin/Downloads/hao.logosc.cn/p/2789.html'
excel_file = 'output_logocom.xlsx'

start_dir = '/Users/deltaqin/Downloads/hao.logosc.cn/p/'
for root, dirs, files in os.walk(start_dir):
    for file in files:
        if file.endswith('.html'):
            html_file = os.path.join(root, file)
            (name, url , div_content) = extract_content_from_html(html_file)
            print("=========" + name + "==============================")
            print(name, url , div_content)
            write_to_excel(name, url , div_content, excel_file)


# content = extract_content_from_html(html_file)
# write_to_excel(content, excel_file)
